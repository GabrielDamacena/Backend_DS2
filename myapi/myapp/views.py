from rest_framework import generics
from .models import Usuario,Ponto
from .serializers import UsuarioSerializer,PontoSerializer
import os
from django.conf import settings
from django.views import View
from rest_framework.response import Response
from rest_framework import status
from .serializers import PhotoUploadSerializer
from deepface import DeepFace
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Alignment, Font




class UsuarioListCreateView(generics.ListCreateAPIView):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

class UsuarioDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

class PontoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

class PontoRegistroView(generics.ListCreateAPIView):
    queryset = Ponto.objects.all()
    serializer_class = PontoSerializer


class ComparePhotoView(generics.CreateAPIView):  # Alterado para CreateAPIView
    serializer_class = PhotoUploadSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Salvar a foto temporária
        photo = serializer.validated_data['photo']
        temp_photo_path = os.path.join(settings.MEDIA_ROOT, 'temp_photos', photo.name)
        os.makedirs(os.path.dirname(temp_photo_path), exist_ok=True)
        
        with open(temp_photo_path, 'wb') as f:
            for chunk in photo.chunks():
                f.write(chunk)

        # Comparar com imagens em media/usuarios
        user_photos_dir = os.path.join(settings.MEDIA_ROOT, 'usuarios')
        if not os.path.exists(user_photos_dir):
            os.makedirs(user_photos_dir)

        comparison_results = []
        for user_photo in os.listdir(user_photos_dir):
            user_photo_path = os.path.join(user_photos_dir, user_photo)
            try:
                # Comparar usando DeepFace
                result = DeepFace.verify(temp_photo_path, user_photo_path)
                comparison_results.append({
                    "user_photo": user_photo,
                    "verified": result['verified'],
                    "distance": result.get('distance'),
                })
            except Exception as e:
                comparison_results.append({
                    "user_photo": user_photo,
                    "error": str(e),
                })

        # Deletar a foto temporária
        os.remove(temp_photo_path)

        # Retornar os resultados
        return Response({"results": comparison_results}, status=status.HTTP_200_OK)
    

class RelatorioPontoPDFView(View):
    def get(self, request, *args, **kwargs):
        # Configurar a resposta como PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_ponto.pdf"'

        # Criar o canvas para desenhar no PDF
        pdf = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        # Título do relatório
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(50, height - 50, "Relatório de Pontos")

        # Cabeçalho da tabela
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, height - 100, "Usuário")
        pdf.drawString(200, height - 100, "Data/Hora")
        pdf.drawString(350, height - 100, "Status")
        pdf.drawString(450, height - 100, "Reconhecido")

        # Linhas de separação
        pdf.line(50, height - 105, 550, height - 105)

        # Dados
        pdf.setFont("Helvetica", 10)
        y = height - 130  # Posição inicial

        pontos = Ponto.objects.all().select_related('user')  # Otimiza as queries
        for ponto in pontos:
            usuario = ponto.user.nome if ponto.user else "Usuário Desconhecido"
            pdf.drawString(50, y, usuario)
            pdf.drawString(200, y, ponto.data_hora.strftime("%d/%m/%Y %H:%M:%S"))
            pdf.drawString(350, y, ponto.get_status_display())
            pdf.drawString(450, y, "Sim" if ponto.reconhecido else "Não")

            y -= 20  # Avança para a próxima linha
            if y < 50:  # Gera uma nova página se necessário
                pdf.showPage()
                pdf.setFont("Helvetica", 10)
                y = height - 50

        pdf.save()
        return response
    

class RelatorioPontoExcelView(View):
    def get(self, request, *args, **kwargs):
        # Criação do arquivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Relatório de Pontos"

        # Título
        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.value = "Relatório de Pontos"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Cabeçalho da tabela
        headers = ['Usuário', 'Data/Hora', 'Status', 'Reconhecido', 'Latitude, Longitude']
        sheet.append(headers)
        for col in sheet.iter_cols(min_row=2, max_row=2, min_col=1, max_col=5):
            for cell in col:
                cell.font = Font(bold=True)

        # Dados da tabela
        pontos = Ponto.objects.all().select_related('user')
        for ponto in pontos:
            usuario = ponto.user.nome if ponto.user else "Usuário Desconhecido"
            reconhecido = "Sim" if ponto.reconhecido else "Não"
            sheet.append([
                usuario,
                ponto.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
                ponto.get_status_display(),
                reconhecido,
                ponto.lat_long
            ])

        # Ajustar largura das colunas
        column_widths = [20, 20, 15, 12, 25]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Resposta HTTP com o arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_ponto.xlsx"'
        workbook.save(response)
        return response